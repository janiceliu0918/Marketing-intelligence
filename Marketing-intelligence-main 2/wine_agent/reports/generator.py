"""Report generator — produces console output, Markdown, and Excel exports."""
import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.columns import Columns
from rich import box
from jinja2 import Template

from wine_agent.models.wine import WineIntelligenceReport, LandedCostBreakdown
from wine_agent.processors.sentiment import aggregate_critic_scores, build_sentiment_summary
from wine_agent.config.settings import config

logger = logging.getLogger(__name__)
console = Console()

# ── Score display helpers ─────────────────────────────────────────────────────

def _score_colour(score: Optional[int]) -> str:
    if score is None:
        return "white"
    if score >= 8:
        return "bright_green"
    if score >= 6:
        return "yellow"
    return "red"


def _rating_bar(score: Optional[float], max_val: float = 100) -> str:
    if score is None:
        return "N/A"
    filled = round((score / max_val) * 15)
    return "█" * filled + "░" * (15 - filled) + f"  {score:.0f}/{max_val:.0f}"


# ── Rich console report ───────────────────────────────────────────────────────

def print_report(report: WineIntelligenceReport) -> None:
    """Print a fully-formatted buyer's brief to the terminal."""
    console.rule(f"[bold wine] BONVIN BUYER'S BRIEF [/bold wine]", style="dark_red")

    # Header
    console.print(Panel(
        f"[bold]{report.wine_name} {report.vintage or ''}[/bold]\n"
        f"Producer: {report.producer}  |  "
        f"Query ID: {report.query_id}  |  "
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        title="[bold red]Wine Intelligence Report[/bold red]",
        border_style="dark_red",
    ))

    # Opportunity Score
    score_colour = _score_colour(report.opportunity_score)
    console.print(
        f"\n[bold]Opportunity Score:[/bold] "
        f"[{score_colour}]{report.opportunity_score or 'N/A'} / 10[/{score_colour}]\n"
    )

    # Executive Summary
    if report.executive_summary:
        console.print(Panel(
            report.executive_summary,
            title="[bold]Executive Summary[/bold]",
            border_style="blue",
        ))

    # Classification
    cls = report.classification
    if cls.appellation or cls.region:
        cls_table = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
        cls_table.add_column("Field", style="dim")
        cls_table.add_column("Value")
        rows = [
            ("Country", cls.country), ("Region", cls.region),
            ("Appellation", cls.appellation), ("Classification", cls.classification),
            ("AOC/VDP Level", cls.aoc_vdp_level), ("Vintage Rating", cls.vintage_rating or ""),
            ("Oak Aging", f"{cls.oak_aging_months} months" if cls.oak_aging_months else ""),
            ("Alcohol", f"{cls.alcohol_pct}%" if cls.alcohol_pct else ""),
        ]
        if cls.grape_varieties:
            blend = ", ".join(f"{k} {v}%" for k, v in cls.grape_varieties.items())
            rows.append(("Blend", blend))
        for field_name, val in rows:
            if val:
                cls_table.add_row(field_name, str(val))
        console.print(Panel(cls_table, title="[bold]Classification & Appellation[/bold]", border_style="green"))

    # Pricing & Cost
    _print_pricing_panel(report)

    # Critic scores
    if report.critic_scores:
        agg = aggregate_critic_scores(report.critic_scores)
        score_table = Table(box=box.SIMPLE, show_header=True, padding=(0, 1))
        score_table.add_column("Critic", style="bold")
        score_table.add_column("Score", justify="center")
        score_table.add_column("Bar")
        for cs in report.critic_scores:
            score_table.add_row(
                cs.critic,
                f"{cs.score:.0f}" if cs.score else "N/A",
                _rating_bar(cs.score),
            )
        score_table.add_section()
        score_table.add_row("[bold]Average[/bold]", f"[bold]{agg['average']}[/bold]",
                            f"[bold]{agg['consensus']}[/bold]")
        console.print(Panel(score_table, title="[bold]Critic Scores[/bold]", border_style="magenta"))

    # Consumer Sentiment
    if report.consumer_sentiment:
        summary = build_sentiment_summary(report.consumer_sentiment)
        console.print(Panel(summary, title="[bold]Consumer Sentiment (Vivino)[/bold]", border_style="cyan"))

    # Vintage Assessment
    if report.vintage_assessment:
        console.print(Panel(report.vintage_assessment, title="[bold]Vintage Assessment[/bold]",
                            border_style="yellow"))

    # Market Positioning
    if report.market_positioning:
        console.print(Panel(report.market_positioning, title="[bold]Market Positioning[/bold]",
                            border_style="bright_blue"))

    # Risk Flags
    if report.risk_flags:
        flag_text = "\n".join(f"[yellow]⚠[/yellow]  {f}" for f in report.risk_flags)
        console.print(Panel(flag_text, title="[bold red]Risk Flags[/bold red]", border_style="red"))

    # Buyer Recommendation
    if report.buyer_recommendation:
        # Detect verdict word
        rec_upper = report.buyer_recommendation.upper()
        if "BUY" in rec_upper and "PASS" not in rec_upper and "HOLD" not in rec_upper:
            verdict_style = "bold bright_green"
        elif "PASS" in rec_upper:
            verdict_style = "bold red"
        else:
            verdict_style = "bold yellow"

        console.print(Panel(
            f"[{verdict_style}]{report.buyer_recommendation}[/{verdict_style}]",
            title="[bold]Buyer Recommendation[/bold]",
            border_style="dark_red",
        ))

    console.rule(style="dark_red")


def _print_pricing_panel(report: WineIntelligenceReport) -> None:
    price_lines = []

    if report.price_benchmarks:
        cad_prices = [p.price_cad for p in report.price_benchmarks if p.price_cad]
        if cad_prices:
            avg = sum(cad_prices) / len(cad_prices)
            price_lines.append(f"Wine-Searcher Global Avg:  CAD {avg:.2f}")

    lc = report.landed_cost
    if lc and lc.total_landed_cad:
        price_lines += [
            "",
            f"Supplier FOB (CAD):        CAD {lc.fob_price_cad:.2f}",
            f"Freight & Insurance:      +CAD {lc.freight_cad + lc.insurance_cad:.2f}",
            f"Import Duty (CETA):       +CAD {lc.import_duty_cad:.2f}",
            f"Federal Excise:           +CAD {lc.federal_excise_cad:.2f}",
            f"LDB Markup (89%):         +CAD {lc.ldb_markup_cad:.2f}",
            f"GST (5%):                 +CAD {lc.gst_cad:.2f}",
            "                           ──────────────",
            f"Estimated Shelf Price:     CAD {lc.suggested_retail_cad:.2f}",
        ]
        if lc.quote_vs_market:
            price_lines.append(f"\nQuote vs Market:  {lc.quote_vs_market}")
        if lc.gross_margin_pct is not None:
            price_lines.append(f"Price Delta:       {lc.gross_margin_pct:+.1f}%")

    if price_lines:
        console.print(Panel(
            "\n".join(price_lines),
            title="[bold]Pricing & BC Landed Cost[/bold]",
            border_style="green",
        ))


# ── Markdown export ───────────────────────────────────────────────────────────

_MD_TEMPLATE = Template("""# Bonvin Buyer's Brief

**Wine**: {{ report.wine_name }} {{ report.vintage or '' }}
**Producer**: {{ report.producer }}
**Generated**: {{ report.generated_at }}
**Query ID**: {{ report.query_id }}

---

## Opportunity Score: {{ report.opportunity_score or 'N/A' }} / 10

## Executive Summary
{{ report.executive_summary }}

## Classification & Appellation
| Field | Value |
|---|---|
| Country | {{ cls.country }} |
| Region | {{ cls.region }} |
| Appellation | {{ cls.appellation }} |
| Classification | {{ cls.classification }} |
| AOC/VDP | {{ cls.aoc_vdp_level }} |
| Vintage Rating | {{ cls.vintage_rating or 'N/A' }} |
{% if cls.grape_varieties %}
| Grape Blend | {{ blend }} |
{% endif %}

## Pricing & BC Landed Cost
{% if lc %}
| Item | Amount |
|---|---|
| Supplier FOB | CAD {{ "%.2f"|format(lc.fob_price_cad) }} |
| Freight & Insurance | CAD {{ "%.2f"|format(lc.freight_cad) }} |
| Import Duty | CAD {{ "%.2f"|format(lc.import_duty_cad) }} |
| Federal Excise | CAD {{ "%.2f"|format(lc.federal_excise_cad) }} |
| LDB Markup (89%) | CAD {{ "%.2f"|format(lc.ldb_markup_cad) }} |
| GST (5%) | CAD {{ "%.2f"|format(lc.gst_cad) }} |
| **Estimated Shelf Price** | **CAD {{ "%.2f"|format(lc.suggested_retail_cad) }}** |
{% if lc.quote_vs_market %}
| Quote vs Market | {{ lc.quote_vs_market }} |
{% endif %}
{% endif %}

{% if market_avg %}
**Wine-Searcher Global Avg**: CAD {{ "%.2f"|format(market_avg) }}
{% endif %}

## Vintage Assessment
{{ report.vintage_assessment }}

## Market Positioning
{{ report.market_positioning }}

## Risk Flags
{% for flag in report.risk_flags %}
- ⚠ {{ flag }}
{% endfor %}

## Buyer Recommendation
{{ report.buyer_recommendation }}

---
*Report generated by Bonvin Wine Market Intelligence Agent*
""")


def export_markdown(report: WineIntelligenceReport, output_dir: Optional[str] = None) -> str:
    """Render and save a Markdown buyer's brief. Returns the file path."""
    cls = report.classification
    lc = report.landed_cost
    blend = ", ".join(f"{k} {v}%" for k, v in cls.grape_varieties.items()) if cls.grape_varieties else ""
    market_avg = None
    if report.price_benchmarks:
        prices = [p.price_cad for p in report.price_benchmarks if p.price_cad]
        market_avg = sum(prices) / len(prices) if prices else None

    md_content = _MD_TEMPLATE.render(
        report=report, cls=cls, lc=lc, blend=blend, market_avg=market_avg
    )

    out_dir = Path(output_dir or config.reports_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_name = report.wine_name.replace(" ", "_").replace("/", "-")
    filename = out_dir / f"brief_{safe_name}_{report.vintage or 'NV'}_{report.query_id}.md"

    with open(filename, "w", encoding="utf-8") as f:
        f.write(md_content)

    logger.info("Markdown report saved: %s", filename)
    return str(filename)


def export_excel(report: WineIntelligenceReport, output_dir: Optional[str] = None) -> str:
    """Export the report as an Excel workbook for Power BI ingestion."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel export")
        return ""

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B0000")

    headers = ["Field", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    lc = report.landed_cost
    cls = report.classification
    rows = [
        ("Wine Name", report.wine_name),
        ("Producer", report.producer),
        ("Vintage", report.vintage),
        ("Query ID", report.query_id),
        ("Generated At", report.generated_at),
        ("Opportunity Score", report.opportunity_score),
        ("Country", cls.country),
        ("Region", cls.region),
        ("Appellation", cls.appellation),
        ("Classification", cls.classification),
        ("Grape Blend", json.dumps(cls.grape_varieties) if cls.grape_varieties else ""),
        ("Vintage Rating", cls.vintage_rating or ""),
        ("Supplier FOB (CAD)", lc.fob_price_cad if lc else ""),
        ("Import Duty (CAD)", lc.import_duty_cad if lc else ""),
        ("LDB Markup (CAD)", lc.ldb_markup_cad if lc else ""),
        ("Estimated Shelf (CAD)", lc.suggested_retail_cad if lc else ""),
        ("Quote vs Market", lc.quote_vs_market if lc else ""),
        ("Gross Margin %", lc.gross_margin_pct if lc else ""),
        ("Buyer Recommendation", report.buyer_recommendation[:500] if report.buyer_recommendation else ""),
    ]
    for r, (field_name, val) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=field_name).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    # ── Sheet 2: Price History ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Prices")
    ws2.append(["Source", "Price (CAD)", "Currency", "Bottle (ml)", "Date"])
    for pp in report.price_benchmarks:
        ws2.append([pp.source, pp.price_cad, pp.currency, pp.bottle_size_ml, pp.as_of_date])

    # ── Sheet 3: Critic Scores ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Critics")
    ws3.append(["Critic", "Score", "Out Of", "Review Date"])
    for cs in report.critic_scores:
        ws3.append([cs.critic, cs.score, cs.score_max, cs.review_date])

    out_dir = Path(output_dir or config.reports_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_name = report.wine_name.replace(" ", "_").replace("/", "-")
    filename = out_dir / f"brief_{safe_name}_{report.vintage or 'NV'}_{report.query_id}.xlsx"
    wb.save(str(filename))
    logger.info("Excel report saved: %s", filename)
    return str(filename)


def export_json(report: WineIntelligenceReport, output_dir: Optional[str] = None) -> str:
    """Dump the report as JSON for downstream consumption."""
    out_dir = Path(output_dir or config.reports_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_name = report.wine_name.replace(" ", "_").replace("/", "-")
    filename = out_dir / f"brief_{safe_name}_{report.vintage or 'NV'}_{report.query_id}.json"

    # dataclass → dict (shallow, handles nested dataclasses)
    import dataclasses
    data = dataclasses.asdict(report)

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

    logger.info("JSON report saved: %s", filename)
    return str(filename)
